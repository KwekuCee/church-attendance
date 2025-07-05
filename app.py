from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, random, qrcode
import json
from flask import send_file
from tempfile import NamedTemporaryFile




app = Flask(__name__)
app.secret_key = 'f4b2d8e9c7a14c6f9b8f14a129c2d654'

USERS_FILE = 'users.json'

FILE_NAME = 'attendance.xlsx'

def get_admin_files():
    username = session.get('username')
    attendance_file = f"{username}_attendance.xlsx"
    members_file = f"{username}_members.xlsx"
    return attendance_file, members_file

def generate_code():
    return f"MKC-{random.randint(100, 999)}"

if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Full Name", "Code", "Invited By", "Service Type", "Date", "Time"])
    wb.save(FILE_NAME)

if not os.path.exists("members.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Full Name", "Code", "Invited By", "Phone"])
    wb.save("members.xlsx")

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = request.get_json()
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    # 🔁 Use the admin's own files
    admin = data.get('admin_username')
    if not admin:
        return jsonify({'status': 'error', 'message': 'Admin username (church) is required'}), 400

    attendance_file = f"{admin}_attendance.xlsx"
    members_file = f"{admin}_members.xlsx"

    # ✅ Create files if missing
    if not os.path.exists(attendance_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Index"
        ws.append(["Full Name", "Code", "Invited By", "Service Type", "Phone", "Date", "Time"])
        wb.save(attendance_file)

    if not os.path.exists(members_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(["Full Name", "Code", "Invited By", "Phone"])
        wb.save(members_file)

    wb_a = load_workbook(attendance_file)
    wb_m = load_workbook(members_file)
    ws_m = wb_m.active

    service_type = data.get('service_type', '').strip()
    sheet_name = f"{service_type} - {today}"

     # ✅ Use or create the right worksheet
    if sheet_name in wb_a.sheetnames:
        ws_a = wb_a[sheet_name]
    else:
        ws_a = wb_a.create_sheet(title=sheet_name)
        ws_a.append(["Full Name", "Code", "Invited By", "Service Type", "Phone", "Date", "Time"])


    # ✅ RETURNING MEMBER
    if 'code' in data and data['code']:
        code = data['code'].strip()
        member = None
        for row in ws_m.iter_rows(min_row=2, values_only=True):
            if row[1] == code:
                member = row
                break

        if not member:
            return jsonify({'status': 'error', 'message': 'Code not found'})

        for row in ws_a.iter_rows(min_row=2, values_only=True):
            if row[1] == code and row[5] == today:
                return jsonify({'status': 'error', 'message': 'Already marked attendance for today.'})

        ws_a.append([
            member[0],  # Full Name
            member[1],  # Code
            member[2],  # Invited By
            service_type,
            member[3],  # Phone
            today,
            now.strftime('%H:%M:%S')
        ])
        wb_a.save(attendance_file)
        return jsonify({
            'status': 'success',
            'code': code,
            'qr_url':f"/static/qr/{code}.png"
        })

    # ✅ NEW MEMBER
    else:
        fullname = data['fullname'].strip()
        phone = data.get('phone', '').strip()
        invited_by = data.get('invited_by', '').strip()
        

        for row in ws_m.iter_rows(min_row=2, values_only=True):
            if row[0].strip().lower() == fullname.lower() and row[3].strip() == phone:
                return jsonify({'status': 'error', 'message': 'This person already has a code. Use returning member option.'})

        code = f"MKC-{random.randint(100, 999)}"

        #Generate QR Code
        qr = qrcode.make(code)
        qr_folder = os.path.join('static', 'qrcodes')
        os.makedirs(qr_folder, exist_ok=True)
        qr_path = os.path.join(qr_folder, f"{code}.png")
        qr.save(qr_path)

        # Save to members file
        ws_m.append([fullname, code, invited_by, phone])
        wb_m.save(members_file)

        # Log attendance
        ws_a.append([
            fullname,
            code,
            invited_by,
            service_type,
            phone,
            today,
            now.strftime('%H:%M:%S')
        ])

        wb_a.save(attendance_file)
        return jsonify({
            'status': 'success',
            'code': code,
            'qr_url':f"/static/qrcodes/{code}.png"
        })
    


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return send_from_directory('.', 'login.html')
    
    username = request.form.get('username')
    password = request.form.get('password')

    if not username or not password:
        return jsonify({'status': 'error', 'message': 'Username and password are required'}), 400

    # Load users from users.json
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            try:
                users = json.load(f)
            except json.JSONDecodeError:
                users = {}
    else:
        users = {}

    user = users.get(username)
    if user and user.get("password") == password:
        session['logged_in'] = True
        session['username'] = username

        # Create admin-specific Excel files if missing
        attendance_file = f"{username}_attendance.xlsx"
        members_file = f"{username}_members.xlsx"

        if not os.path.exists(attendance_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Full Name", "Code", "Invited By", "Service Type", "Phone", "Date", "Time"])
            wb.save(attendance_file)

        if not os.path.exists(members_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Full Name", "Code", "Invited By", "Phone"])
            wb.save(members_file)

        return jsonify({'status': 'success'}), 200

    return jsonify({'status': 'error', 'message': 'Invalid login credentials'}), 401


@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'GET':
        return send_from_directory(',', 'register.html')
    
    username = request.form.get('username')
    password = request.form.get('password')
    church_name = request.form.get('church')

    if not username or not password:
        return "Missing isername or password or church name", 400
    
    users = {}
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            try:
                users = json.load(f)
            except json.JSONDecodeError:
                users = {}


    if username in users:
        return "Username already exists", 400

    users[username] = {
    "password": password,
    "church": church_name
    }

 #Store as plain text (simplified)

    with open(USERS_FILE, 'w') as f:
        json.dump(users, f)

    return redirect ('/login')        

@app.route('/churches')
def churches():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            users = json.load(f)
        return jsonify([
            {"admin": username, "church": details.get("church", "Unnamed Church")}
            for username, details in users.items()
        ])
    return jsonify([])


@app.route('/admin')
def admin():
    if not session.get('logged_in'):
        return redirect('/login')
    return send_from_directory('.', 'admin.html')

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return redirect('/login')

@app.route('/records')
def records():
    if not session.get('logged_in'):
        return redirect('/login')

    admin = session.get('username')
    attendance_file = f"{admin}_attendance.xlsx"

    if not os.path.exists(attendance_file):
        return jsonify([])

    wb = load_workbook(attendance_file)
    all_records = []

    for sheet in wb.sheetnames:
        if sheet == "Index":
            continue  # skip the default sheet

        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_records.append(row)

    return jsonify(all_records)


@app.route('/export')
def export_attendance():
    if not session.get('logged_in'):
        return redirect('/login')

    admin = session.get('username')
    attendance_file = f"{admin}_attendance.xlsx"
    service_type = request.args.get('service_type')
    date = request.args.get('date')

    if not service_type or not date:
        return jsonify({'message': 'Missing service or date'}), 400

    sheet_name = f"{service_type} - {date}"

    # Debugging info
    print(f"[Export] Looking for file: {attendance_file}")
    print(f"[Export] Looking for sheet: {sheet_name}")

    if not os.path.exists(attendance_file):
        return jsonify({'message': f'File {attendance_file} not found'}), 404

    wb = load_workbook(attendance_file)
    if sheet_name not in wb.sheetnames:
        return jsonify({'message': f'Sheet {sheet_name} not found in {attendance_file}', 'sheets': wb.sheetnames}), 404

    from tempfile import NamedTemporaryFile
    tmp = NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    for row in wb[sheet_name].iter_rows(values_only=True):
        new_ws.append(row)

    new_wb.save(tmp_path)
    return send_file(tmp_path, as_attachment=True, download_name=f"{sheet_name}.xlsx")


@app.route('/<path:path>')
def static_file(path):
    return send_from_directory('.', path)

@app.route('/session-status')
def session_status():
    return jsonify({'logged_in': session.get('logged_in', False)})

@app.route('/delete-record', methods=['POST'])
def delete_record():
    if not session.get('logged_in'):
        return redirect('/login')

    data = request.get_json()
    fullname = data.get('fullname', '').strip().lower()
    code = data.get('code')
    date = data.get('date')
    time = data.get('time')

    admin = session.get('username')
    attendance_file = f"{admin}_attendance.xlsx"

    if not os.path.exists(attendance_file):
        return jsonify({'status': 'error', 'message': 'Attendance file not found'}), 404

    wb = load_workbook(attendance_file)

    # Search all sheets except "Index"
    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue

        ws = wb[sheet_name]

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_name = str(row[0]).strip().lower() if row[0] else ""
            row_code = row[1]
            row_date = row[5]
            row_time = row[6]

            if row_name == fullname and row_code == code and row_date == date and row_time == time:
                ws.delete_rows(idx)
                wb.save(attendance_file)
                return jsonify({'status': 'success'})

    return jsonify({'status': 'error', 'message': 'Record not found'}), 404
    
@app.route('/members')
def get_members():
    if not session.get('logged_in'):
        return redirect('/login')

    admin = session.get('username')
    members_file = f"{admin}_members.xlsx"

    if not os.path.exists(members_file):
        return jsonify([])

    wb = load_workbook(members_file)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append({
            "fullname": row[0],
            "code": row[1],
            "invited_by": row[2],
            "phone": row[3]
        })
    return jsonify(data)

@app.route('/members/update', methods=['POST'])
def update_member():
    if not session.get('logged_in'):
        return redirect('/login')

    data = request.get_json()
    code = data.get("code")
    fullname = data.get("fullname").strip()
    invited_by = data.get("invited_by").strip()
    phone = data.get("phone").strip()

    admin = session.get('username')
    members_file = f"{admin}_members.xlsx"

    if not os.path.exists(members_file):
        return jsonify({'status': 'error', 'message': 'Members file not found'}), 404

    wb = load_workbook(members_file)
    ws = wb.active

    updated = False
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(code):
            row[0].value = fullname
            row[2].value = invited_by
            row[3].value = phone
            updated = True
            break

    if updated:
        wb.save(members_file)
        return jsonify({'status': 'success'})
    else:
        return jsonify({'status': 'error', 'message': 'Member not found'}), 404

@app.route('/scan')
def scan_qr_code():
    return send_from_directory('.', 'scan.html')

@app.route('/qr-check-in', methods=['POST'])
def qr_check_in():
    data = request.get_json()
    code = data.get('code')
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    if not code:
        return jsonify({'status': 'error', 'messsage': 'No code provided'}), 400
    
    #Search through all member files
    users = {}
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            users = json.load(f)

    matched = None
    for  username in users:
        members_file = f"{username}_members.xlsx"
        attendance_file = f"{username}_attenndance.xlsx"
        
        if not os.path.exists(members_file):
            continue

        wb_m = load_workbook(members_file)
        ws_m = wb_m.active

        for row in ws_m.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == str(code):
                matched = {
                    "fullname": row[0],
                    "invited_by": row[2],
                    "phone": row [3],
                    "admin": username
                }
                break
            if matched:
                break

        if not matched:
            return jsonify({'status': 'error', 'message': 'Code not found in any church'}), 404
        
        #Now write attendance to that admin's attendance file
        attendace_file = f"{matched['admin']}_attendance.xlsx"
        service_type = "QR Check-In"

        wb_a = load_workbook(attendance_file)
        #Use or create service-specific sheet
        sheet_name = f"{service_type} - {today}"
        if sheet_name not in wb_a.sheetnames:
            wb_a.create_sheet(title=sheet_name)
            wb_a[sheet_name].append(["Full Name", "Code", "Invited By", "Service Type", "Phone", "Date", "Time"])

            ws = wb_a[sheet_name]

        #Avoid duplicate
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == code and row[5] == today:
                return jsonify({'status': 'error', 'message': "Already marked today"}), 400
            
        attendance_row = [
            matched.get('fullname', ''),
            code,
            matched.get('invited_by', ''),
            service_type,
            matched.get('phone', ''),
            today,
            now.strftime('%H:%M:%S')
        ]

        #Ensure it matches the column count in the header
        if len(ws[1]) == len(attendance_row):
            ws.append(attendance_row)
        else:
            return jsonify({'status': 'error', 'message': 'Sheet format mismatch'}), 500
        
    return jsonify({'status': 'success'})

@app.route('/scanner')
def scanner():
    return send_from_directory('.', 'scanner.html')

@app.route('/qr-submit', methods=['POST'])
def qr_submit():
    data = request.get_json()
    code = data.get('code')
    # Look up member by code and log attendance...
    # return jsonify({ status: 'success', name: 'John Doe' })


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)